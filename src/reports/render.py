# Stock Analysis AI Agent - Multi-Exchange Stock Analysis Platform
# Copyright (C) 2026 B.Vignesh Kumar (Bravetux) <ic19939@gmail.com>
# Licensed under GNU GPL v3 or later.

"""Deterministic StockThesis -> Markdown rendering.

Replaces the brittle "LLM builds markdown, then a second LLM re-parses it into
a table" flow in pages/1_Analysis.py. Given a StockThesis, always produce the
same Markdown. Given a list[StockThesis], always produce the same consolidated
table.
"""

from __future__ import annotations

from io import BytesIO
from typing import Iterable

from src.agents.evidence import Evidence, ResearchPlan, Signal, StockThesis


_SIGNAL_EMOJI = {
    Signal.BULLISH: "🟢 BULLISH",
    Signal.BEARISH: "🔴 BEARISH",
    Signal.NEUTRAL: "⚪ NEUTRAL",
    Signal.INCONCLUSIVE: "❔ INCONCLUSIVE",
}


def _fmt_price(v: float | None) -> str:
    if v is None:
        return "—"
    try:
        return f"{float(v):,.2f}"
    except (TypeError, ValueError):
        return str(v)


def _fmt_pct(v: float | None) -> str:
    if v is None:
        return "—"
    try:
        return f"{float(v) * 100:.1f}%"
    except (TypeError, ValueError):
        return str(v)


def render_thesis_markdown(
    thesis: StockThesis,
    *,
    plan: ResearchPlan | None = None,
    evidence: Iterable[Evidence] | None = None,
    include_citations: bool = True,
) -> str:
    """Produce a readable Markdown report from a StockThesis.

    The output is deliberately machine-friendly too: every field the batch
    consolidator needs can be extracted via direct attribute access on the
    StockThesis, not regex on the markdown.
    """
    lines: list[str] = []
    lines.append(f"# Stock Analysis — {thesis.ticker} ({thesis.exchange})")
    lines.append("")
    lines.append(f"*Generated {thesis.generated_at.strftime('%Y-%m-%d %H:%M UTC')}*")
    lines.append("")
    lines.append(f"**Signal:** {_SIGNAL_EMOJI.get(thesis.signal, thesis.signal.value)}  ")
    lines.append(f"**Conviction:** {_fmt_pct(thesis.conviction)}  ")
    lines.append(f"**Headline:** {thesis.headline}")
    lines.append("")

    # Scenarios ---------------------------------------------------------------
    lines.append("## Scenarios")
    lines.append("")
    lines.append("| Case | Probability | Price Target | Horizon (d) | Catalysts | Invalidators |")
    lines.append("|---|---|---|---|---|---|")
    for sc in thesis.scenarios:
        lines.append(
            "| {name} | {prob} | {tgt} | {hz} | {cats} | {invs} |".format(
                name=sc.name,
                prob=_fmt_pct(sc.probability),
                tgt=_fmt_price(sc.price_target),
                hz=sc.time_horizon_days,
                cats="<br>".join(sc.catalysts) or "—",
                invs="<br>".join(sc.invalidators) or "—",
            )
        )
    lines.append("")

    # Key levels --------------------------------------------------------------
    if thesis.key_levels:
        lines.append("## Key Levels")
        lines.append("")
        lines.append("| Level | Value |")
        lines.append("|---|---|")
        for k in ("current_price", "support", "resistance", "dma_200", "vwap", "fib_618"):
            if k in thesis.key_levels:
                lines.append(f"| {k} | {_fmt_price(thesis.key_levels[k])} |")
        # Any extras the synthesizer added that aren't in the canonical list.
        for k, v in thesis.key_levels.items():
            if k not in {"current_price", "support", "resistance", "dma_200", "vwap", "fib_618"}:
                lines.append(f"| {k} | {_fmt_price(v)} |")
        lines.append("")

    # Contradictions resolved -------------------------------------------------
    if thesis.contradictions_resolved:
        lines.append("## Contradictions Resolved")
        lines.append("")
        for c in thesis.contradictions_resolved:
            lines.append(f"- {c}")
        lines.append("")

    # Top evidence (citations) ------------------------------------------------
    if include_citations and thesis.top_evidence:
        lines.append("## Top Evidence")
        lines.append("")
        lines.append("| # | Thread | Signal | Conf. | Claim | Source |")
        lines.append("|---|---|---|---|---|---|")
        for i, ev in enumerate(thesis.top_evidence, start=1):
            lines.append(
                "| {i} | {t} | {s} | {c} | {claim} | {src} |".format(
                    i=i,
                    t=ev.thread_id,
                    s=ev.signal.value,
                    c=f"{ev.confidence:.2f}",
                    claim=ev.claim.replace("|", "\\|"),
                    src=ev.source_tool or "—",
                )
            )
        lines.append("")

    # Data quality ------------------------------------------------------------
    if thesis.data_quality_flags:
        lines.append("## Data Quality Notes")
        lines.append("")
        for f in thesis.data_quality_flags:
            lines.append(f"- {f}")
        lines.append("")

    # Research plan / framing -------------------------------------------------
    if plan is not None:
        lines.append("## Research Plan")
        lines.append("")
        lines.append(f"**Framing:** {plan.framing}")
        lines.append("")
        lines.append("| Thread | Priority | Objective | Budget |")
        lines.append("|---|---|---|---|")
        for t in plan.threads:
            lines.append(f"| {t.thread_id} | {t.priority} | {t.objective} | {t.budget_tool_calls} |")
        lines.append("")

    # Evidence appendix -------------------------------------------------------
    if evidence is not None:
        ev_list = list(evidence)
        if ev_list:
            lines.append("## Evidence Appendix")
            lines.append("")
            by_thread: dict[str, list[Evidence]] = {}
            for ev in ev_list:
                by_thread.setdefault(ev.thread_id, []).append(ev)
            for tid, items in by_thread.items():
                lines.append(f"### {tid}")
                lines.append("")
                for ev in items:
                    lines.append(
                        f"- **[{ev.signal.value}, conf {ev.confidence:.2f}]** {ev.claim} "
                        f"*(source: {ev.source_tool or '—'})*"
                    )
                    if ev.caveats:
                        for cv in ev.caveats:
                            lines.append(f"  - caveat: {cv}")
                lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("*Informational only — not investment advice.*")

    return "\n".join(lines)


# ----------------------------------------------------------------------------
# Batch consolidated table
# ----------------------------------------------------------------------------


def consolidated_rows(theses: list[StockThesis]) -> tuple[list[str], list[list[str]]]:
    """Return (headers, rows) ready to drop into a DataFrame or XLSX."""
    headers = [
        "S.No",
        "Ticker",
        "Exchange",
        "Signal",
        "Conviction",
        "Current Price",
        "Base Target",
        "Bull Target",
        "Bear Target",
        "Support",
        "Resistance",
        "200DMA",
        "Headline",
        "Data Flags",
    ]
    rows: list[list[str]] = []
    for i, t in enumerate(theses, start=1):
        sc_by_name = {s.name: s for s in t.scenarios}
        base = sc_by_name.get("base")
        bull = sc_by_name.get("bull")
        bear = sc_by_name.get("bear")
        rows.append([
            str(i),
            t.ticker,
            t.exchange,
            t.signal.value,
            f"{t.conviction:.2f}",
            _fmt_price(t.key_levels.get("current_price")),
            _fmt_price(base.price_target if base else None),
            _fmt_price(bull.price_target if bull else None),
            _fmt_price(bear.price_target if bear else None),
            _fmt_price(t.key_levels.get("support")),
            _fmt_price(t.key_levels.get("resistance")),
            _fmt_price(t.key_levels.get("dma_200")),
            t.headline.replace("\n", " "),
            "; ".join(t.data_quality_flags) if t.data_quality_flags else "",
        ])
    return headers, rows


def consolidated_markdown_table(theses: list[StockThesis]) -> str:
    headers, rows = consolidated_rows(theses)
    out: list[str] = []
    out.append("| " + " | ".join(headers) + " |")
    out.append("|" + "|".join(["---"] * len(headers)) + "|")
    for r in rows:
        out.append("| " + " | ".join(cell.replace("|", "\\|") for cell in r) + " |")
    return "\n".join(out)


def consolidated_xlsx_bytes(theses: list[StockThesis]) -> bytes:
    """Build an XLSX buffer directly from StockThesis list — no markdown round-trip."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback: build from pandas if openpyxl isn't directly importable.
        import pandas as pd
        headers, rows = consolidated_rows(theses)
        df = pd.DataFrame(rows, columns=headers)
        buf = BytesIO()
        df.to_excel(buf, index=False, engine="openpyxl")
        return buf.getvalue()

    headers, rows = consolidated_rows(theses)
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E5A88")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for r in rows:
        ws.append(r)

    # Auto-size columns (rough)
    for col_idx, header in enumerate(headers, start=1):
        max_len = max(
            [len(header)] + [len(str(r[col_idx - 1])) for r in rows if col_idx <= len(r)]
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
